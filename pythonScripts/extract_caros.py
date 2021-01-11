from os import path
from openpyxl import load_workbook
from Bio import SeqIO
import re 

xls_file = "Sup Table Peptide data.xlsx"
acessions_list_out_prefix = "CAROS_accs"
cathacyc_file = "Catro_mRNA_update.tfa"
output_file = "Catro_mRNA_update.tfa_annotated"


def get_caros_list_from_xls(xls_file, output_prefix):
    #Pulls caros list from the xls file and store per in seperate files (1 per sheet)
    wb = load_workbook(xls_file, data_only=True)
    for sheet in wb: 
        acession_list = []
        for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
            #each instance is a tuple (due to implementation probably). Get the first element of the tuple, 
            #check if it is different to "none" and add to a list. Most row[0] will be equal to none. 
            if row[0] != None: 
                acession_list.append(row[0])

        with open("{}_{}.txt".format(acessions_list_out_prefix, sheet.title), "w") as fh: 
            for name in acession_list: 
                fh.write(name)
                fh.write("\n")
        
        fh2 = open("temp.xls_parsed.ok", "w")
        fh2.close()
        

def parse_db_cathacyc_file(cathacyc_file): 
    #this function parses the cathacyc file into a dictionary, with the keys being the third word in the description
    id2_sequence_map = {}
    db_cathacyc =  SeqIO.parse(cathacyc_file, "fasta")
    for seq in db_cathacyc:
        seq_id = seq.description.split(" ")[2] 
        id2_sequence_map[seq_id] = (seq, [])
    return id2_sequence_map



def annotate_withsample(files, id2_sequence_map):
    #annotates in the second position of the tuple the samples in which each gene appear 
    file_i = 0  
    for file in files: 
        with open(file, "r") as fh: 
            for cnt, line in enumerate(fh):
                if "Contig" in line: 
                    m = re.search("_(Contig[0-9]*)_", line)
                    map_key = m.group(1)
                if "Caros" in line: 
                    m = re.search("_(Caros[0-9]*.1)_", line)
                    map_key = m.group(1)
                if "Locus" in line: 
                    m = re.search("_(Locus_[0-9]*_Transcript_[0-9]*)_", line)
                    map_key = m.group(1)
                
                if file_i == 0:
                    id2_sequence_map[map_key][1].append("VAC_1")
                if file_i == 1:
                    id2_sequence_map[map_key][1].append("VAC_2")
                if file_i == 2:
                    id2_sequence_map[map_key][1].append("Ton_1")
                if file_i == 3:
                    id2_sequence_map[map_key][1].append("Ton_2")    

        file_i += 1 
    return id2_sequence_map    

def extract_annotated_caros(id2_sequence_map): 
    #modifies the header and keeps in a list the acession hits 
    annotated_caros_list = []
    for key, value in id2_sequence_map.items():
        #check if there is an empty list, ie the sequences does not appear in any of the samples 
        if value[1] != []: 
            annot_text = "[{}]".format(" ".join(value[1]))
            value[0].description = value[0].description + " " + annot_text
            annotated_caros_list.append(value[0])
    return annotated_caros_list

def main():
    if not path.exists("temp.xls_parsed.ok"): 
        get_caros_list_from_xls(xls_file, acessions_list_out_prefix)
    id2_sequence_map = parse_db_cathacyc_file(cathacyc_file)
    #Change these as needed
    files_list = ["CAROS_accs_ic120607_1_G_Vac#1.txt", "CAROS_accs_ic120607_2_G_Vac#2.txt", "CAROS_accs_ic120607_3_G_Ton#1.txt", "CAROS_accs_ic120607_4_G_Ton#2.txt"]
    id2_2_sequence_map = annotate_withsample(files_list, id2_sequence_map)
    annotated_caros_list = extract_annotated_caros(id2_2_sequence_map)
    
    SeqIO.write(annotated_caros_list, output_file, "fasta")


main()