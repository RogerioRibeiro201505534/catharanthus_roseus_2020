#use this python script to extract all the peptides and compare with the trypsin cleaved sites 

from os import path
from openpyxl import load_workbook
import csv

xls_file = "Sup Table Peptide data.xlsx"
new_file_list = ["peptides_#vac1.txt", "peptides_#vac2.txt", "peptides_#ton1.txt", "peptides_#ton2.txt"]

def get_peptides_list(xls_file, new_file_list):
    #Pulls caros list from the xls file and store each sample in seperate files (1 per sheet)
    print("Loading file .... ")
    wb = load_workbook(xls_file, data_only=True)
    print("File loaded!")
    k = 0
    for sheet in wb: 
        peptides_list = []
        for row in sheet.iter_rows(min_row=1, max_col=3, values_only=True):
            #check if it is different to "none" and add to a list (rows with proteins and not peptides)
            if row[0] == None: 
                peptides_list.append(row[2])

        peptides_list = list(set(peptides_list))
        with open(new_file_list[k], "w") as fh:
            for pep in peptides_list: 
                fh.write(pep)
                fh.write("\n")
        k = k + 1

    

get_peptides_list(xls_file, new_file_list)