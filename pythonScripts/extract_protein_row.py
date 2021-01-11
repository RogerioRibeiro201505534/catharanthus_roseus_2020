from os import path
from openpyxl import load_workbook
from Bio import SeqIO
import re 
import csv

xls_file = "Sup Table Peptide data.xlsx"
cathacyc_file = "Catro_mRNA_update.tfa_annotated"
proteome_table_list = ["proteome_table_ic120607_1_G_Vac#1_not_treated.tsv", "proteome_table_ic120607_2_G_Vac#2_not_treated.tsv", "proteome_table_ic120607_3_G_Ton#1_not_treated.tsv", "proteome_table_ic120607_4_G_Ton#2_not_treated.tsv"]
new_file_list = ["proteome_table_Vac#1_not_treated_v2.tsv", "proteome_table_Vac#2_not_treated_v2.tsv", "proteome_table_Ton#1_not_treated_v2.tsv", "proteome_table_Ton#2_not_treated_v2.tsv"]

def get_caros_list_from_xls(xls_file):
    #Pulls caros list from the xls file and store each sample in seperate files (1 per sheet)
    print("Loading file .... ")
    wb = load_workbook(xls_file, data_only=True)
    print("File loaded!")
    for sheet in wb: 
        row_list = []
        for row in sheet.iter_rows(min_row=1, max_col=12, values_only=True):
            #check if it is different to "none" and add to a list (rows with proteins and not peptides)
            if row[0] != None: 
                row_list.append(row)
            
        with open("proteome_table_{}_not_treated.tsv".format(sheet.title), "w") as fh: 
            for row2print in row_list: 
                for element in row2print: 
                    fh.write(str(element))
                    fh.write("\t")
                fh.write("\n")


        
def parse_db_cathacyc_file(cathacyc_file): 
    #this function parses the cathacyc acession numbers into a dictionary, with the keys being the third word in the description
    id2caros_map = {}
    db_cathacyc =  SeqIO.parse(cathacyc_file, "fasta")
    for seq in db_cathacyc:
        seq_id = seq.description.split(" ")
        id2caros_map[seq_id[2]] = seq_id[0]
    return id2caros_map

def modify_proteome_table(proteome_table_list, id2caros_map, new_file_list): 
    #Modifies the first collum in the file to the correspondent caros acession number 
    i = 0
    for table in proteome_table_list: 
        header_flag = True
        print(table)
        with open(table, newline="\n") as fh: 
            table_reader = csv.reader(fh, delimiter = "\t")
            for row in table_reader: 
                #if it is the header run this code 
                if header_flag == True: 
                    newfh = open(new_file_list[i], "w") 
                    newfh.write("\t".join(row))
                    newfh.write("\n")
                    newfh.close()
                    header_flag = False
                #else run this 
                else: 
                    if "Contig" in row[0]: 
                        m = re.search("_(Contig[0-9]*)_", row[0])
                        map_key = m.group(1)
                    if "Caros" in row[0]: 
                        m = re.search("_(Caros[0-9]*.1)_", row[0])
                        map_key = m.group(1)
                    if "Locus" in row[0]: 
                        m = re.search("_(Locus_[0-9]*_Transcript_[0-9]*)_", row[0])
                        map_key = m.group(1)
                    
                    row[0] = id2caros_map[map_key]
                    newfh = open(new_file_list[i], "a")
                    newfh.write("\t".join(row))
                    newfh.write("\n")
                    newfh.close()
        i += 1         

#get_caros_list_from_xls(xls_file)
id2caros_map = parse_db_cathacyc_file(cathacyc_file)
modify_proteome_table(proteome_table_list, id2caros_map, new_file_list)

